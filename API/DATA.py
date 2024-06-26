import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
import warnings

warnings.simplefilter(action="ignore", category=FutureWarning)


class GestionDatos:
    def __init__(self):
        self.nombre_archivo = "registros.xlsx"
        self.crear_dataframes()
        self.cargar_dataframes()
        self.guardar_dataframes()

    def crear_dataframes(self):
        self.columnas_clientes = ["Cedula", "Nombre", "Telefono"]
        self.columnas_productos = [
            "Referencia",
            "Codigo de barras",
            "Marca",
            "Precio de adquisicion",
            "Precio venta",
            "Unidades actuales",
            "Producto disponible",
            "Fecha",
        ]
        self.columnas_metodo_pago = ["ID", "Metodo de pago"]
        self.columnas_venta_productos = [
            "ID venta",
            "Cedula",
            "Cliente",
            "Producto",
            "Fecha",
            "Cantidad",
            "Subtotal",
            "ID_MetodoPago",
        ]
        self.columnas_venta_servicios = [
            "ID venta",
            "Cedula",
            "Cliente",
            "Servicio",
            "Fecha",
            "Cantidad",
            "Subtotal",
            "ID_MetodoPago",
        ]
        self.columnas_servicios = ["ID servicio", "Nombre Servicio", "Costo"]
        self.columnas_reservas_servicios = [
            "ID Reserva",
            "Cliente",
            "Servicio",
            "Fecha Reserva",
            "Fecha Servicio",
        ]
        self.columnas_facturas = [
            "Cedula",
            "Cliente",
            "Fecha",
            "SubtotalProductos",
            "SubtotalServicios",
            "Total",
            "Metodo de pago",
        ]
        self.columnas_usuarios = ["ID usuario", "usuario", "contraseña", "Rol ID"]
        self.columnas_roles = ["ID", "Rol"]

        self.clientes = pd.DataFrame(columns=self.columnas_clientes)
        self.productos = pd.DataFrame(columns=self.columnas_productos)
        datos_metodo_pago = {"ID": [1, 2], "Metodo de pago": ["efectivo", "tarjeta"]}
        self.metodo_pago = pd.DataFrame(
            datos_metodo_pago, columns=self.columnas_metodo_pago
        )
        self.venta_productos = pd.DataFrame(columns=self.columnas_venta_productos)
        self.venta_servicios = pd.DataFrame(columns=self.columnas_venta_servicios)
        self.servicios = pd.DataFrame(columns=self.columnas_servicios)
        self.reservas_servicios = pd.DataFrame(columns=self.columnas_reservas_servicios)
        self.facturas = pd.DataFrame(columns=self.columnas_facturas)
        self.usuarios = pd.DataFrame(columns=self.columnas_usuarios)
        datos_roles = {"ID": [1, 2, 3], "Rol": ["soporte", "admin", "caja"]}
        self.roles = pd.DataFrame(datos_roles, columns=self.columnas_roles)

        # Crear un usuario de soporte predeterminado
        usuario_soporte = pd.DataFrame(
            [
                [1, "soporte", "hola", 1],
                [2, "cajero", "cajero", 3],
                [3, "admin", "admin", 2],
            ],
            columns=self.columnas_usuarios,
        )
        self.usuarios = pd.concat([self.usuarios, usuario_soporte], ignore_index=True)

    def cargar_dataframes(self):
        try:
            archivo = pd.ExcelFile(self.nombre_archivo)
            self.clientes = pd.read_excel(archivo, sheet_name="Clientes")
            self.productos = pd.read_excel(archivo, sheet_name="Productos")
            self.metodo_pago = pd.read_excel(archivo, sheet_name="Metodo de pago")
            self.venta_productos = pd.read_excel(archivo, sheet_name="VentaProductos")
            self.venta_servicios = pd.read_excel(archivo, sheet_name="VentaServicios")
            self.servicios = pd.read_excel(archivo, sheet_name="Servicios")
            self.reservas_servicios = pd.read_excel(
                archivo, sheet_name="ReservasServicios"
            )
            self.facturas = pd.read_excel(archivo, sheet_name="Facturas")
            self.usuarios = pd.read_excel(archivo, sheet_name="Usuarios")
            self.roles = pd.read_excel(archivo, sheet_name="Roles")
        except FileNotFoundError:
            print(
                f"Archivo {self.nombre_archivo} no encontrado. Se crearán nuevos DataFrames."
            )
        except Exception as e:
            print(f"Error al cargar los DataFrames desde el archivo Excel: {e}")

    def guardar_dataframes(self):
        try:
            with pd.ExcelWriter(self.nombre_archivo) as writer:
                # Guardar todos los DataFrames en el archivo Excel
                self.clientes.to_excel(writer, sheet_name="Clientes", index=False)
                self.productos.to_excel(writer, sheet_name="Productos", index=False)
                self.metodo_pago.to_excel(
                    writer, sheet_name="Metodo de pago", index=False
                )
                self.venta_productos.to_excel(
                    writer, sheet_name="VentaProductos", index=False
                )
                self.venta_servicios.to_excel(
                    writer, sheet_name="VentaServicios", index=False
                )
                self.servicios.to_excel(writer, sheet_name="Servicios", index=False)
                self.reservas_servicios.to_excel(
                    writer, sheet_name="ReservasServicios", index=False
                )
                self.facturas.to_excel(writer, sheet_name="Facturas", index=False)
                self.usuarios.to_excel(writer, sheet_name="Usuarios", index=False)
                self.roles.to_excel(writer, sheet_name="Roles", index=False)
            self.ajustar_columnas_excel()
        except Exception as e:
            print(f"Error al guardar los dataframes en el archivo Excel: {e}")

    def ajustar_columnas_excel(self):
        try:
            archivo = load_workbook(self.nombre_archivo)
            for hoja in archivo.sheetnames:
                hoja_actual = archivo[hoja]
                for columnas in hoja_actual.columns:
                    max_length = 0
                    columna = [celda for celda in columnas]
                    for celda in columna:
                        try:
                            if celda.value is not None:
                                max_length = max(max_length, len(str(celda.value)))
                        except Exception as e:
                            print(f"Error al procesar la celda {celda.coordinate}: {e}")
                    ajusta_ancho = max_length + 2
                    columna_letra = get_column_letter(columna[0].column)
                    hoja_actual.column_dimensions[columna_letra].width = ajusta_ancho
            archivo.save(self.nombre_archivo)
        except Exception as e:
            print(f"Error al ajustar las columnas en el archivo Excel: {e}")

    def verificar_admin_existente(self):
        return "Admin" in self.roles["Rol"].values

    def verificar_stock_no_negativo(self):
        return all(self.productos["Unidades actuales"] >= 0)

    def realizar_verificaciones(self):
        if not self.verificar_admin_existente():
            print("Advertencia: No hay usuarios administradores en el sistema.")
        if not self.verificar_stock_no_negativo():
            print("Error: El stock no puede ser negativo en el inventario.")

    # Clientes
    def agregar_cliente(self, cedula, nombre, telefono):
        nuevo_cliente = pd.DataFrame(
            [[cedula, nombre, telefono]], columns=self.columnas_clientes
        )
        self.clientes = pd.concat([self.clientes, nuevo_cliente], ignore_index=True)
        self.guardar_dataframes()

    def buscar_cliente(self, cedula):
        cliente = self.clientes[self.clientes["Cedula"] == cedula]
        if not cliente.empty:
            return cliente
        else:
            print(f"Cliente con cedula {cedula} no encontrado.")
            return None

    def eliminar_clientes(self, cedula):
        cliente = self.clientes[self.clientes["Cedula"] == cedula]
        if not cliente.empty:
            self.clientes = self.clientes[self.clientes["Cedula"] != cedula]
            self.guardar_dataframes()
            return True
        else:
            return False

    def actualizar_cliente(self, cedula, nuevos_datos):
        cliente = self.clientes[self.clientes["Cedula"] == cedula]
        if not cliente.empty:
            for key, value in nuevos_datos.items():
                if key in self.clientes.columns:
                    self.clientes.loc[self.clientes["Cedula"] == cedula, key] = value
            self.guardar_dataframes()
            return True
        else:
            return False

    # Productos
    def agregar_producto(
        self,
        referencia,
        codigo_barras,
        marca,
        precio_adquisicion,
        precio_venta,
        unidades_actuales,
    ):
        fecha = datetime.now().strftime("%Y-%m-%d") 
        disponible = int(unidades_actuales) > 0
        nuevo_producto = pd.DataFrame(
            [
                [
                    referencia,
                    codigo_barras,
                    marca,
                    precio_adquisicion,
                    precio_venta,
                    unidades_actuales,
                    disponible,
                    fecha,
                ]
            ],
            columns=self.columnas_productos,
        )
        # self.productos = pd.concat([self.productos, nuevo_producto], ignore_index=True)
        self.guardar_dataframes()

        # Especificar los tipos de datos de las columnas
        tipos_de_datos = {
            "Referencia": str,
            "Codigo de barras": str,
            "Marca": str,
            "Precio de adquisicion": float,
            "Precio venta": float,
            "Unidades actuales": int,
            "Producto disponible": bool,
            "Fecha": str,
        }
        nuevo_producto = nuevo_producto.astype(tipos_de_datos)

        # Comprobar si el DataFrame 'nuevo_producto' no está vacío y no contiene solo valores NA
        if not nuevo_producto.empty and not nuevo_producto.isna().all().all():
            self.productos = pd.concat(
                [self.productos, nuevo_producto], ignore_index=True
            )

        self.guardar_dataframes()

    def buscar_producto(self, codigo):
        codigo = int(codigo)
        producto = self.productos[self.productos["Codigo de barras"] == codigo]
        if not producto.empty:
            return True
        else:
            return False

    def verificar_disponibilidad(self, codigo):
        producto = self.productos[self.productos["Codigo de barras"] == codigo]
        if not producto.empty:
            if producto.iloc[0]["Unidades actuales"] > 0:
                print(f"El producto con referencia {codigo} está disponible.")
            else:
                print(f"El producto con referecodigoBarrascia {codigo} no está disponible.")
        else:
            print(f"Producto con la referencia: {codigo} no ha sido encontrado")

    def actualizar_producto(self, codigoBarras, nuevos_datos):
        producto = self.productos[self.productos["Codigo de barras"] == codigoBarras]
        if not producto.empty:
            for key, value in nuevos_datos.items():
                if key in self.productos.columns:
                    self.productos.loc[
                        self.productos["Codigo de barras"] == codigoBarras, key
                    ] = value
            self.guardar_dataframes()
            #print(f"Producto con codigoBarras {codigoBarras} ha sido actualizado.")
        else:
            #print(f"Producto con codigoBarras {codigoBarras} no encontrado.")
            None    
    
    def descontinuar_producto(self,codigo):
        producto = self.productos[self.productos["Codigo de barras"] == codigo]
        if not producto.empty:
            self.productos = self.productos[self.productos["Codigo de barras"] != codigo]
            self.guardar_dataframes()
            return True
        else:
            return False

    # Venta de productos y servicios
    def agregar_venta_servicio(
        self,
        id_servicio,
        cedula,
        cliente,
        nombre_servicio,
        cantidad,
        subtotal,
        ID_MetodoPago,
    ):
        fecha = datetime.now().strftime("%d/%m/%Y %H:%M")
        nueva_venta_servicio = pd.DataFrame(
            [
                [
                    id_servicio,
                    cedula,
                    cliente,
                    nombre_servicio,
                    fecha,
                    cantidad,
                    subtotal,
                    ID_MetodoPago,
                ]
            ],
            columns=self.columnas_venta_servicios,
        )
        self.venta_servicios = pd.concat(
            [self.venta_servicios, nueva_venta_servicio], ignore_index=True
        )
        self.guardar_dataframes()

    def agregar_venta_producto(
        self, id_venta, cedula, cliente, producto, cantidad, subtotal, ID_MetodoPago
    ):
        fecha = datetime.now().strftime("%d/%m/%Y %H:%M")
        nueva_venta_producto = pd.DataFrame(
            [
                [
                    id_venta,
                    cedula,
                    cliente,
                    producto,
                    fecha,
                    cantidad,
                    subtotal,
                    ID_MetodoPago,
                ]
            ],
            columns=self.columnas_venta_productos,
        )
        self.venta_productos = pd.concat(
            [self.venta_productos, nueva_venta_producto], ignore_index=True
        )
        self.guardar_dataframes()

    def actualizar_venta(self, id_venta, nuevos_datos, tipo_venta="producto"):
        if tipo_venta == "producto":
            venta = self.venta_productos[self.venta_productos["ID venta"] == id_venta]
            if not venta.empty:
                for key, value in nuevos_datos.items():
                    if key in self.venta_productos.columns:
                        self.venta_productos.loc[
                            self.venta_productos["ID venta"] == id_venta, key
                        ] = value
                self.guardar_dataframes()
                print(f"Venta de producto con ID {id_venta} actualizada.")
            else:
                print(f"Venta de producto con ID {id_venta} no encontrada.")
        elif tipo_venta == "servicio":
            venta = self.venta_servicios[self.venta_servicios["ID venta"] == id_venta]
            if not venta.empty:
                for key, value in nuevos_datos.items():
                    if key in self.venta_servicios.columns:
                        self.venta_servicios.loc[
                            self.venta_servicios["ID venta"] == id_venta, key
                        ] = value
                self.guardar_dataframes()
                print(f"Venta de servicio con ID {id_venta} actualizada.")
            else:
                print(f"Venta de servicio con ID {id_venta} no encontrada.")

    def eliminar_venta(self, id_venta, tipo_venta="producto"):
        if tipo_venta == "producto":
            venta = self.venta_productos[self.venta_productos["ID venta"] == id_venta]
            if not venta.empty:
                self.venta_productos = self.venta_productos[
                    self.venta_productos["ID venta"] != id_venta
                ]
                self.guardar_dataframes()
                print(f"El producto con el ID: {id_venta} ha sido eliminado")
            else:
                print(f"El producto con el ID: {id_venta} no ha sido encontrado")
        elif tipo_venta == "servicio":
            venta = self.venta_servicios[self.venta_servicios["ID venta"] == id_venta]
            if not venta.empty:
                self.venta_servicios = self.venta_servicios[
                    self.venta_servicios["ID venta"] != id_venta
                ]
                self.guardar_dataframes()
                print(f"El servicio con el ID: {id_venta} ha sido eliminado")
            else:
                print(f"El servicio con el ID: {id_venta} no ha sido encontrado")

    # Servicios
    def agregar_servicio(self, id_servicio, nombre_servicio, costo):
        nuevo_servicio = pd.DataFrame(
            [[id_servicio, nombre_servicio, costo]], columns=self.columnas_servicios
        )
        self.servicios = pd.concat([self.servicios, nuevo_servicio], ignore_index=True)
        self.guardar_dataframes()

    def actualizar_servicio(self, id_servicio, nuevos_datos):
        servicio = self.servicios[self.servicios['ID servicio'] == id_servicio]
        if not servicio.empty:
            for key, value in nuevos_datos.items():
                if key in self.servicios.columns:
                    self.servicios.loc[self.servicios['ID servicio'] == id_servicio, key] = value
            self.guardar_dataframes()

    def buscar_servicio(self, id_servicio):
        servicio = self.servicios[self.servicios["ID servicio"] == id_servicio]
        if not servicio.empty:
            return servicio
        else:
            print(f"Servicio con ID {id_servicio} no encontrado.")
            return None

    def eliminar_servicio(self, id_servicio):
        servicio = self.servicios[self.servicios["ID servicio"] == id_servicio]
        if not servicio.empty:
            self.servicios = self.servicios[
                self.servicios["ID servicio"] != id_servicio
            ]
            self.guardar_dataframes()
            print(f"Servicio con ID {id_servicio} eliminado.")
        else:
            print(f"Servicio con ID {id_servicio} no encontrado.")

    # Reserva de Servicios
    def reservar_servicio(self, id_reserva, cliente, servicio, fecha_servicio):
        fecha_reserva = datetime.now().strftime("%d/%m/%Y %H:%M")
        nueva_reserva = pd.DataFrame(
            [[id_reserva, cliente, servicio, fecha_reserva, fecha_servicio]],
            columns=self.columnas_reservas_servicios,
        )
        self.reservas_servicios = pd.concat(
            [self.reservas_servicios, nueva_reserva], ignore_index=True
        )
        self.guardar_dataframes()

    def cancelar_reserva_servicio(self, id_reserva):
        reserva = self.reservas_servicios[
            self.reservas_servicios["ID Reserva"] == id_reserva
        ]
        if not reserva.empty:
            self.reservas_servicios = self.reservas_servicios[
                self.reservas_servicios["ID Reserva"] != id_reserva
            ]
            self.guardar_dataframes()
        else:
            return None

    def buscar_reserva_por_servicio(self, id_servicio):
        reservas = self.reservas_servicios[
            self.reservas_servicios["Servicio"] == id_servicio
        ]
        if not reservas.empty:
            return reservas
        else:
            return None

    # Usuarios
    def agregar_usuario(self, id_usuario, usuario, contraseña, rol_id):
        nuevo_usuario = pd.DataFrame(
            [[id_usuario, usuario, contraseña, rol_id]], columns=self.columnas_usuarios
        )

        # Insertar debajo del usuario predeterminado
        indice_usuario_predeterminado = self.usuarios[
            self.usuarios["usuario"] == "soporte"
        ].index

        if not indice_usuario_predeterminado.empty:
            indice_insertar = indice_usuario_predeterminado[0] + 1
            self.usuarios = pd.concat(
                [
                    self.usuarios.iloc[:indice_insertar],
                    nuevo_usuario,
                    self.usuarios.iloc[indice_insertar:],
                ],
                ignore_index=True,
            )
        else:
            self.usuarios = pd.concat([self.usuarios, nuevo_usuario], ignore_index=True)

        self.guardar_dataframes()

    def buscar_usuario(self, usuario):
        usuarios = self.usuarios[self.usuarios["usuario"] == usuario]
        if not usuarios.empty:
            return usuarios
        else:
            print(f"El usuario {usuario} no fue encontrado")
            return None

    def modificar_usuario(self, usuario, nuevos_datos):
        usuarios = self.usuarios[self.usuarios["usuario"] == usuario]
        if not usuarios.empty:
            for key, value in nuevos_datos.items():
                if key in self.usuarios.columns:
                    self.usuarios.loc[self.usuarios["usuario"] == usuario, key] = value
            self.guardar_dataframes()
            print(f"Usuario {usuario} modificado correctamente.")
        else:
            print(f"El usuario {usuario} no ha sido encontrado.")

    # Facturas
    def conversion_factura(self, cedula, ID_MetodoPago):
        fecha = datetime.now().strftime("%d/%m/%Y %H:%M")
        try:
            cliente_servicios = self.venta_servicios[
                (self.venta_servicios["Cedula"] == cedula)
                & (self.venta_servicios["Fecha"] == fecha)
            ]
            cliente_productos = self.venta_productos[
                (self.venta_productos["Cedula"] == cedula)
                & (self.venta_productos["Fecha"] == fecha)
            ]

            if not cliente_servicios.empty:
                nombre_cliente = cliente_servicios["Cliente"].iloc[0]
            elif not cliente_productos.empty:
                nombre_cliente = cliente_productos["Cliente"].iloc[0]
            else:
                print(
                    f"No se encontraron ventas para el cliente con cédula {cedula} en la fecha {fecha}."
                )
                return

            total_venta_servicios = 0
            total_venta_productos = 0
            if not cliente_servicios.empty:
                total_venta_servicios = cliente_servicios["Subtotal"].sum()
            if not cliente_productos.empty:
                total_venta_productos = cliente_productos["Subtotal"].sum()
            total = total_venta_servicios + total_venta_productos

            nueva_factura = pd.DataFrame(
                [
                    [
                        cedula,
                        nombre_cliente,
                        fecha,
                        total_venta_productos,
                        total_venta_servicios,
                        total,
                        ID_MetodoPago,
                    ]
                ],
                columns=self.columnas_facturas,
            )
            self.facturas = pd.concat([self.facturas, nueva_factura], ignore_index=True)
            self.guardar_dataframes()
        except Exception as e:
            print(f"Error al generar la factura: {e}")
