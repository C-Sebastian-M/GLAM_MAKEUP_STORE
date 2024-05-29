import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime


class GestionDatos:
    def __init__(self, nombre_archivo):
        self.nombre_archivo = nombre_archivo
        self.crear_dataframes()
        self.guardar_dataframes()

    def crear_dataframes(self):
        self.columnas_clientes = ["Cedula", "Nombre", "Telefono"]
        self.columnas_productos = ["Referencia", "Codigo de barras", "Marca", "Precio de adquisicion", "Precio venta",
                                   "Unidades actuales"]
        self.columnas_venta_productos = ["ID venta", "Cliente", "Producto", "Fecha", "Cantidad", "Subtotal"]
        self.columnas_venta_servicios = ["ID venta", "Cliente", "Servicio", "Fecha", "Costo"]
        self.columnas_servicios = ["ID servicio", "Nombre Servicio", "Costo"]
        self.columnas_contraseñas = ["Usuario", "Contraseña", "Rol"]
        self.columnas_inventario = ["Producto", "Stock"]
        self.columnas_reservas_servicios = ["ID Reserva", "Cliente", "Servicio", "Fecha Reserva", "Fecha Servicio"]

        self.clientes = pd.DataFrame(columns=self.columnas_clientes)
        self.productos = pd.DataFrame(columns=self.columnas_productos)
        self.venta_productos = pd.DataFrame(columns=self.columnas_venta_productos)
        self.venta_servicios = pd.DataFrame(columns=self.columnas_venta_servicios)
        self.servicios = pd.DataFrame(columns=self.columnas_servicios)
        self.contraseñas = pd.DataFrame(columns=self.columnas_contraseñas)
        self.inventario = pd.DataFrame(columns=self.columnas_inventario)
        self.reservas_servicios = pd.DataFrame(columns=self.columnas_reservas_servicios)

    def guardar_dataframes(self):
        try:
            with pd.ExcelWriter(self.nombre_archivo) as writer:
                self.clientes.to_excel(writer, sheet_name='Clientes', index=False)
                self.productos.to_excel(writer, sheet_name='Productos', index=False)
                self.venta_productos.to_excel(writer, sheet_name='VentaProductos', index=False)
                self.venta_servicios.to_excel(writer, sheet_name='VentaServicios', index=False)
                self.servicios.to_excel(writer, sheet_name='Servicios', index=False)
                self.contraseñas.to_excel(writer, sheet_name='Contraseñas', index=False)
                self.inventario.to_excel(writer, sheet_name='Inventario', index=False)
                self.reservas_servicios.to_excel(writer, sheet_name='ReservasServicios', index=False)
            self.ajustar_columnas_excel()
        except Exception as e:
            print(f"Error al guardar los dataframes en el archivo Excel: {e}")

    def ajustar_columnas_excel(self):
        try:
            archivo = load_workbook(self.nombre_archivo)
            for hoja in archivo.sheetnames:
                hoja_actual = archivo[hoja]
                for columnas in hoja_actual.columns:
                    max_length = 0
                    columna = [celda for celda in columnas]
                    for celda in columna:
                        try:
                            if celda.value is not None:
                                max_length = max(max_length, len(str(celda.value)))
                        except Exception as e:
                            print(f"Error al procesar la celda {celda.coordinate}: {e}")
                    ajusta_ancho = max_length + 2
                    columna_letra = get_column_letter(columna[0].column)
                    hoja_actual.column_dimensions[columna_letra].width = ajusta_ancho
            archivo.save(self.nombre_archivo)
        except Exception as e:
            print(f"Error al ajustar las columnas en el archivo Excel: {e}")

    def verificar_admin_existente(self):
        return 'Admin' in self.contraseñas['Rol'].values

    def verificar_stock_no_negativo(self):
        return all(self.inventario['Stock'] >= 0)

    def realizar_verificaciones(self):
        if not self.verificar_admin_existente():
            print("Advertencia: No hay usuarios administradores en el sistema.")
        if not self.verificar_stock_no_negativo():
            print("Error: El stock no puede ser negativo en el inventario.")

    # Clientes
    def agregar_cliente(self, cedula, nombre, telefono):
        nuevo_cliente = pd.DataFrame([[cedula, nombre, telefono]], columns=self.columnas_clientes)
        self.clientes = pd.concat([self.clientes, nuevo_cliente], ignore_index=True)
        self.guardar_dataframes()

    def buscar_cliente(self, cedula):
        cliente = self.clientes[self.clientes['Cedula'] == cedula]
        if not cliente.empty:
            return cliente
        else:
            print(f"Cliente con cedula {cedula} no encontrado.")
            return None

    def eliminar_clientes(self, cedula):
        cliente = self.clientes[self.clientes["Cedula"] == cedula]
        if not cliente.empty:
            self.clientes = self.clientes[self.clientes["Cedula"] != cedula]
            self.guardar_dataframes()
            print(f"Cliente con cedula: {cedula} ha sido eliminado")
        else:
            print(f"El cliente con cedula {cedula} no ha sido encontrado.")

    def actualizar_cliente(self, cedula, nuevos_datos):
        cliente = self.clientes[self.clientes['Cedula'] == cedula]
        if not cliente.empty:
            for key, value in nuevos_datos.items():
                if key in self.clientes.columns:
                    self.clientes.loc[self.clientes['Cedula'] == cedula, key] = value
            self.guardar_dataframes()
            print(f"Cliente con cedula {cedula} ha sido actualizado.")
        else:
            print(f"Cliente con cedula {cedula} no encontrado.")

    # Productos
    def agregar_producto(self, referencia, codigo_barras, marca, precio_adquisicion, precio_venta, unidades_actuales):
        nuevo_producto = pd.DataFrame(
            [[referencia, codigo_barras, marca, precio_adquisicion, precio_venta, unidades_actuales]],
            columns=self.columnas_productos)
        self.productos = pd.concat([self.productos, nuevo_producto], ignore_index=True)
        self.guardar_dataframes()

    def buscar_producto(self, referencia):
        producto = self.productos[self.productos["Referencia"] == referencia]
        if not producto.empty:
            return producto
        else:
            print(f"Producto con la referencia: {referencia} no ha sido encontrado")
            return None

    def eliminar_producto(self, referencia):
        producto = self.productos[self.productos['Referencia'] == referencia]
        if not producto.empty:
            self.productos = self.productos[self.productos['Referencia'] != referencia]
            self.guardar_dataframes()
            print(f"Producto con referencia {referencia} eliminado.")
        else:
            print(f"Producto con referencia {referencia} no encontrado.")

    def actualizar_producto(self, referencia, nuevos_datos):
        producto = self.productos[self.productos['Referencia'] == referencia]
        if not producto.empty:
            for key, value in nuevos_datos.items():
                if key in self.productos.columns:
                    self.productos.loc[self.productos['Referencia'] == referencia, key] = value
            self.guardar_dataframes()
            print(f"Producto con referencia {referencia} ha sido actualizado.")
        else:
            print(f"Producto con referencia {referencia} no encontrado.")

    # Venta de productos y servicios
    def agregar_venta_servicio(self, id_venta, cliente, servicio, fecha, costo):
        nueva_venta_servicio = pd.DataFrame([[id_venta, cliente, servicio, fecha, costo]],
                                            columns=self.columnas_venta_servicios)
        self.venta_servicios = pd.concat([self.venta_servicios, nueva_venta_servicio], ignore_index=True)
        self.guardar_dataframes()

    def agregar_venta_producto(self, id_venta, cliente, producto, fecha, cantidad, subtotal):
        nueva_venta_producto = pd.DataFrame([[id_venta, cliente, producto, fecha, cantidad, subtotal]],
                                            columns=self.columnas_venta_productos)
        self.venta_productos = pd.concat([self.venta_productos, nueva_venta_producto], ignore_index=True)
        self.guardar_dataframes()

    def actualizar_venta(self, id_venta, nuevos_datos, tipo_venta='producto'):
        if tipo_venta == 'producto':
            venta = self.venta_productos[self.venta_productos['ID venta'] == id_venta]
            if not venta.empty:
                self.venta_productos.update(nuevos_datos)
                self.guardar_dataframes()
                print(f"Venta de producto con ID {id_venta} actualizada.")
            else:
                print(f"Venta de producto con ID {id_venta} no encontrada.")
        elif tipo_venta == 'servicio':
            venta = self.venta_servicios[self.venta_servicios['ID venta'] == id_venta]
            if not venta.empty:
                self.venta_servicios.update(nuevos_datos)
                self.guardar_dataframes()
                print(f"Venta de servicio con ID {id_venta} actualizada.")
            else:
                print(f"Venta de servicio con ID {id_venta} no encontrada.")

    def eliminar_venta(self, id_venta, tipo_venta="producto"):
        if tipo_venta == "producto":
            venta = self.venta_productos[self.venta_productos["ID venta"] == id_venta]
            if not venta.empty:
                self.venta_productos = self.venta_productos[self.venta_productos["ID venta"] != id_venta]
                self.guardar_dataframes()
                print(f"El producto con el ID: {id_venta} ha sido eliminado")
            else:
                print(f"El producto con el ID: {id_venta} no ha sido encontrado")
        elif tipo_venta == "servicio":
            venta = self.venta_servicios[self.venta_servicios["ID venta"] == id_venta]
            if not venta.empty:
                self.venta_servicios = self.venta_servicios[self.venta_servicios["ID venta"] != id_venta]
                self.guardar_dataframes()
                print(f"El servicio con el ID: {id_venta} ha sido eliminado")
            else:
                print(f"El servicio con el ID: {id_venta} no ha sido encontrado")

    # Servicios
    def agregar_servicio(self, id_servicio, nombre_servicio, costo):
        nuevo_servicio = pd.DataFrame([[id_servicio, nombre_servicio, costo]], columns=self.columnas_servicios)
        self.servicios = pd.concat([self.servicios, nuevo_servicio], ignore_index=True)
        self.guardar_dataframes()

    def actualizar_servicio(self, id_servicio, nuevos_datos):
        servicio = self.servicios[self.servicios["ID servicio"] == id_servicio]
        if not servicio.empty:
            self.servicios.update(nuevos_datos)
            self.guardar_dataframes()
            print(f"El servicio con el ID: {id_servicio} ha sido actualizado")
        else:
            print(f"El servicio con el ID: {id_servicio} no pudo ser encontrado")

    def buscar_servicio(self, id_servicio):
        servicio = self.servicios[self.servicios['ID servicio'] == id_servicio]
        if not servicio.empty:
            return servicio
        else:
            print(f"Servicio con ID {id_servicio} no encontrado.")
            return None

    def eliminar_servicio(self, id_servicio):
        servicio = self.servicios[self.servicios["ID servicio"] == id_servicio]
        if not servicio.empty:
            self.servicios = self.servicios[self.servicios["ID servicio"] != id_servicio]
            self.guardar_dataframes()
            print(f"Servicio con ID {id_servicio} eliminado.")
        else:
            print(f"Servicio con ID {id_servicio} no encontrado.")

    # Reserva de Servicios
    def reservar_servicio(self, id_reserva, cliente, servicio, fecha_servicio):
        fecha_reserva = datetime.now().strftime("%d/%m/%Y %H:%M")
        nueva_reserva = pd.DataFrame([[id_reserva, cliente, servicio, fecha_reserva, fecha_servicio]],
                                     columns=self.columnas_reservas_servicios)
        self.reservas_servicios = pd.concat([self.reservas_servicios, nueva_reserva], ignore_index=True)
        self.guardar_dataframes()

    def cancelar_reserva_servicio(self, id_reserva):
        reserva = self.reservas_servicios[self.reservas_servicios['ID Reserva'] == id_reserva]
        if not reserva.empty:
            self.reservas_servicios = self.reservas_servicios[self.reservas_servicios['ID Reserva'] != id_reserva]
            self.guardar_dataframes()
        else:
            return None

    def buscar_reserva_por_servicio(self, id_servicio):
        reservas = self.reservas_servicios[self.reservas_servicios['Servicio'] == id_servicio]
        if not reservas.empty:
            return reservas
        else:
            return None

    # Contraseñas
    def agregar_contraseña(self, usuario, contraseña, rol):
        nueva_contraseña = pd.DataFrame([[usuario, contraseña, rol]], columns=self.columnas_contraseñas)
        self.contraseñas = pd.concat([self.contraseñas, nueva_contraseña], ignore_index=True)
        self.guardar_dataframes()

    def buscar_usuario(self, usuario):
        usuarios = self.contraseñas[self.contraseñas["Usuario"] == usuario]
        if not usuarios.empty:
            return usuarios
        else:
            print(f"El usuario {usuario} no fue encontrado")
            return None

    def eliminar_usuario(self, usuario):
        usuarios = self.contraseñas[self.contraseñas["Usuario"] == usuario]
        if not usuarios.empty:
            self.contraseñas = self.contraseñas[self.contraseñas["Usuario"] != usuario]
            self.guardar_dataframes()
        else:
            print(f"El usuario {usuario} no ha sido encontrado")


x = GestionDatos('datos.xlsx')
x.crear_dataframes()
x.reservar_servicio(12, "mgiue", 345, "28/05/2024 14:30")
x.guardar_dataframes()
